# -*- coding: utf-8 -*-
"""
Created on Wed Nov 10 16:43:56 2021

@author: iant

CONVERT ALL PPT AND PPTX TO PDF AND MAKE A TABLE ADDING TO THE WEBSITE

1. PUT ALL PPT, PPTX OR PDF PRESENTATIONS IN A DIRECTORY
1. MAKE A NEW PRESENTATIONS.XLXS CONTAINING SECTION HEADINGS, PRESENTATION NAMES, AUTHOR AND PRESENTATION FILENAME INC EXTENSION
2. RUN SCRIPT TO MAKE TABLE FOR WEBSITE
"""

import os
import posixpath
import sys

from openpyxl import load_workbook
# from openpyxl.styles import PatternFill


from tools.file.ppt_pptx_to_pdf import ppt_pptx_to_pdf


meeting_dir = r"C:\Users\iant\Documents\DOCUMENTS\SWTs"

header_dict = {
    6:"SWT6: CSL, Liege (20-22 April 2015)",
    7:"SWT7: Salobrena, near Granada (20-21 October 2015)",
    8:"SWT8: IAPS, Rome (11-13 May 2016)",
    9:"SWT9: Goddard SFC, Washington DC (7-9 Dec 2016)",
    10:"SWT10: BIRA-IASB, Brussels (29-31 May 2017)",
    11:"SWT11: IDR, Madrid (13-14 November 2017)",
    12:"SWT12: Royal Astronomical Society, London (20-22 August 2018)",
    13:"SWT13: ASI, Rome (28-31 January 2019)",
    14:"SWT14: IAA, Granada (22-23 May 2019)",
    15:"SWT15: Space Research Centre, Wroclaw (12-13 September 2019)",
    16:"SWT16 (Joint with ACS+MAVEN): Sorbonne University, Paris (17-21 February 2020)",
    17:"SWT17: Virtual Meeting (16-17 June 2020)",
    18:"SWT18 Virtual Meeting (7-9 December 2020)",
    19:"SWT19: Virtual Meeting (24-26 March 2021)",
    20:"SWT20: BIRA + Virtual (20-22 October 2021)",
    21:"SWT21: Virtual Meeting (29-31 March 2022)",
    }

# swt_numbers = [20, 19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6]
swt_numbers = [21]
dir_list = os.listdir(meeting_dir)


h = ""
# h = "<html><head></head><body>\n"

for swt_number in swt_numbers:
    print(swt_number)


    swt_dirname = [v for v in dir_list if "SWT%i" %(swt_number) in v][0]
    
    
    swt_dir = os.path.join(meeting_dir, swt_dirname)
    href_dir = posixpath.join(r"ProjectDir/meetings", swt_dirname)
    
    xlsx_filepath = os.path.join(swt_dir, "presentations.xlsx")
    
    
    
    #open spreadsheet
    if not os.path.exists(xlsx_filepath):
        print("%s does not exist" %xlsx_filepath)
        sys.exit()
    
    wb = load_workbook(xlsx_filepath, data_only=True)
    sheets = wb.sheetnames
    Sheet1 = wb["Sheet1"]
    
    #count number of entries in summary file
    data_in = []
    for row_number in range(1000):
        data_row = []
        #check if summary file row contains data
        value = Sheet1.cell(row_number+1, 1).value
        if value != None:
            data_row.append(value)
            
            value = Sheet1.cell(row_number+1, 2).value
            if value != None:
                data_row.append(value)
    
                value = Sheet1.cell(row_number+1, 3).value
                if value != None:
                    data_row.append(value)
    
        data_in.append(data_row)
        # n_rows_xlsx = sum(contains_data)
    
    data_len = [i for i, v in enumerate(data_in) if len(v) != 0][-1] #index of last non-zero element
    data = [v for i, v in enumerate(data_in) if i <= data_len]
        
    """convert ppts to pdf"""
    for data_row in data:
        if len(data_row) == 3:
            filename = data_row[2]
            extension = os.path.splitext(filename)[1]
    
            filepath_in = os.path.join(swt_dir, filename)
            
            day_found = ""
            if os.path.exists(filepath_in):
                print(filename, "found")
            
            else:
                #try other folders
                
                days = [v for v in os.listdir(swt_dir) if "Day" in v]
                for day in days:
                    filepath_in = os.path.join(swt_dir, day, filename)
                    # print("Trying path %s" %filepath_in)
                    if os.path.exists(filepath_in):
                        print(filename, "found in folder", day)
                        day_found = day
                if day_found != "":
                    filepath_in = os.path.join(swt_dir, day_found, filename)
                else:
                    print("Error: %s not found" %filename)
                    sys.exit()
            
            if extension in ["ppt", ".pptx"]:
                filepath_out = filepath_in.replace(extension, ".pdf")
                
                if not os.path.exists(filepath_out):
                    ppt_pptx_to_pdf(filepath_in, filepath_out)
                    
            else:
                filepath_out = filepath_in
            
            if day_found != "":
                data_row[2] = posixpath.join(day_found, os.path.basename(filepath_out))
            else:
                data_row[2] = os.path.basename(filepath_out)
    
    
    h += "<h1>%s</h1>\n" %header_dict[swt_number]
    
    h += "<table border='1'>\n"
    h += "<tr>\n   <th>Title</th><th>Presenter</th><th>PDF</th>\n</tr>\n"
    
    for data_row in data:
        h += "<tr>\n"
        
        if len(data_row) == 0:
            h += "   <td><br></td>\n"
        
        for i, column in enumerate(data_row):
            if i == 2:
                href_path = posixpath.join(href_dir, column)
                h += "   <td><a href='%s'>%s</a></td>\n" %(href_path, os.path.basename(column))
            else:
                if len(data_row) == 1:  
                    h += "   <td><b>%s</b></td>\n" %column
                else:
                    h += "   <td>%s</td>\n" %column
    
            if i == 1 and len(data_row) == 2: #if no presentation has been provided
                h += "   <td>No presentation provided</td>\n"
    
    
        h += "</tr>\n"
    
    h += "</table>\n"
    h += "<br>\n"
    h += "<br>\n"
# h += "</body></html>\n"
    
with open("swt_presentations.html", "w", encoding="utf-8") as f:
    for i, line in enumerate(h):
        f.write(line)